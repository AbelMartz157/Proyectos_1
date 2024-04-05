import sqlite3
import tkinter as tk
import subprocess
import sys
from tkinter import ttk, messagebox, PhotoImage
from datetime import date, timedelta, datetime
from tkcalendar import DateEntry
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
import locale
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

#Valores iniciales******************************************************************************************************************************************************************************
conexion = sqlite3.connect('Breton industrial.db')  # Cambia el nombre del archivo de la base de datos si es diferente
cursor = conexion.cursor()

#Funciones**************************************************************************************************************************************************************************************
def ver_HrsExtra(event=None):
    query1 = '''
        SELECT
            Empleado,
            SUM(CASE WHEN Horario = 'Extra Matutino' THEN HorasExtra ELSE 0 END) AS TotalHorasExtra1,
            SUM(CASE WHEN Horario = 'Extra Vespertino' THEN HorasExtra ELSE 0 END) AS TotalHorasExtra2
        FROM 'Datos Metalmecanica'
        WHERE Fecha = ?
        GROUP BY Empleado
        HAVING TotalHorasExtra1 > 0 OR TotalHorasExtra2 > 0
    '''
    # Ejecutar la consulta con el valor de la fecha
    cursor.execute(query1, (date_fecha.get_date().strftime('%d/%m/%Y'),))
    # Obtener los resultados
    resultados = cursor.fetchall()
    # Limpiar el Treeview antes de insertar nuevos datos
    for item in HrExtraTotal.get_children():
        HrExtraTotal.delete(item)
    # Insertar los resultados en el Treeview y aplicar etiquetas de color
    for row in resultados:
        HrExtraTotal.insert("", "end", text=row[0], values=(row[1], row[2]))
        HrExtraTotal.item(HrExtraTotal.get_children()[-1], tags=("grande"))

    query2 = '''
        SELECT Empleado, Horario, OT, Proyecto, Ensamble, Operacion, HorasExtra
        FROM 'Datos Metalmecanica'
        WHERE Fecha = ? AND (Horario = 'Extra Matutino' OR Horario = 'Extra Vespertino')
        ORDER BY Empleado, Horario
    '''
    # Ejecutar la consulta con el valor de la fecha
    cursor.execute(query2, (date_fecha.get_date().strftime('%d/%m/%Y'),))
    # Obtener los resultados
    resultados = cursor.fetchall()
    # Limpiar el Treeview antes de insertar nuevos datos
    for item in HrExtra.get_children():
        HrExtra.delete(item)
    # Insertar los resultados en el Treeview y aplicar etiquetas de color
    for row in resultados:
        HrExtra.insert("", "end", text=row[0], values=(row[1], row[2], row[3], row[4], row[5], row[6]))
        HrExtra.item(HrExtra.get_children()[-1], tags=("grande"))

def exportar_excel():
    # Obtener los encabezados del primer Treeview
    encabezados_hr_extra_total = ['Nombre', 'Extra Matutino', 'Hrs Matutino', 'Extra Vespertino', 'Hrs Vespertino', 'Total Horas']
    # Obtener los datos del primer Treeview
    datos_hr_extra_total = [[HrExtraTotal.item(row_id, "text")] + [HrExtraTotal.set(row_id, column) for column in HrExtraTotal["columns"]] for row_id in HrExtraTotal.get_children()]
    # Obtener los encabezados del segundo Treeview
    encabezados_hr_extra = ['Nombre', 'Horario', 'OT', 'Proyecto', 'Ensamble', 'Operacion', 'HorasE']
    # Obtener los datos del segundo Treeview
    datos_hr_extra = [[HrExtra.item(row_id, "text")] + [HrExtra.set(row_id, column) for column in HrExtra["columns"]] for row_id in HrExtra.get_children()]

    # Combinar los datos de ambos Treeviews con un espacio de un renglón
    #datos_combinados = datos_hr_extra_total + [[]] + datos_hr_extra + [[], [], [], [], ["", "", "Firma de responsable:", "________________________________________________"]]
    # Crear un DataFrame de pandas
    #df = pd.DataFrame(datos_combinados[1:], columns=datos_combinados[0])

    # Obtener la fecha y hora actual
    fecha_hora_actual = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Crear el nombre del archivo con la fecha y hora
    nombre_archivo = f"Reporte_Tiempo_Extra_{fecha_hora_actual}.xlsx"

    # Crear un archivo de Excel con openpyxl
    workbook = openpyxl.Workbook()
    hoja_trabajo = workbook.active
    hoja_trabajo.title = 'Reporte TE'
    
    # Agregar una imagen
    imagen = Image('Breton_industrial2.png')  # Reemplaza 'ruta_de_la_imagen.png' con la ruta real de tu imagen
    hoja_trabajo.add_image(imagen, 'A1')  # Puedes ajustar 'A1' según la celda donde deseas colocar la imagen

    # Obtener los encabezados y agregarlos a la hoja de trabajo
    #encabezados = datos_combinados[0]
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')  # Configurar la localización en español
    hoja_trabajo.append(['', '', '', 'Reporte de horas extra'])
    hoja_trabajo.append(['', '', '', '', '', 'Fecha', date_fecha.get_date().strftime("%d/%m/%Y")])
    hoja_trabajo.append(['', '', '', '', '', 'Dia', date_fecha.get_date().strftime("%A")])
    hoja_trabajo.append([])
    
    hoja_trabajo.append(encabezados_hr_extra_total)
    # Obtener los datos y agregarlos a la hoja de trabajo
    for fila in datos_hr_extra_total:
        valor_fila_1 = float(fila[1]) if fila[1] else 0.0
        valor_fila_2 = float(fila[2]) if fila[2] else 0.0

        # Calcular la hora final para las columnas "6am a..." y "5pm a..."
        if datetime.strptime(date_fecha.get(), "%d/%m/%Y").weekday() > 4: #Sabado, Domingo
            hora_iniciomat = datetime.strptime("07:00", "%H:%M")
        else:
            hora_iniciomat = datetime.strptime("06:00", "%H:%M")    
        hora_finalmat = hora_iniciomat + timedelta(hours=valor_fila_1)

        if datetime.strptime(date_fecha.get(), "%d/%m/%Y").weekday() == 4: #Viernes
            hora_iniciovesp = datetime.strptime("15:00", "%H:%M")
        else:
            hora_iniciovesp = datetime.strptime("17:00", "%H:%M")
        hora_finalvesp = hora_iniciovesp + timedelta(hours=valor_fila_2)

        # Formatear las columnas "6am a..." y "5pm a..." en formato de hora
        if valor_fila_1 >= 6:
            #print("Sabado, Domingo")
            extra_6am = f"07:00am a {hora_finalmat.strftime('%I:%M %p')}pm"
        else:
            extra_6am = f"06:00am a {hora_finalmat.strftime('%I:%M %p')}am"
        
        if datetime.strptime(date_fecha.get(), "%d/%m/%Y").weekday() == 4:
            #print("Es viernes")
            extra_5pm = f"03:00pm a {hora_finalvesp.strftime('%I:%M %p')}pm"
        else:
            extra_5pm = f"05:00pm a {hora_finalvesp.strftime('%I:%M %p')}pm"

        # Crear una nueva lista con el formato deseado
        extra_vesp = fila[:1] + [extra_6am] + [''] + [extra_5pm] #+ fila[3:]
        hoja_trabajo.append(extra_vesp)

    hoja_trabajo.append([])
    hoja_trabajo.append(["Actividades"])

    hoja_trabajo.append(encabezados_hr_extra)
    for fila in datos_hr_extra:
        hoja_trabajo.append(fila)
  
    for i in range(3):
        hoja_trabajo.append([])

    hoja_trabajo.append(["", "", "Firma de responsable:"])

    # Configurar los bordes
    borde = Border(left=Side(border_style='thin', color='000000'),  # Borde izquierdo
            right=Side(border_style='thin', color='000000'),  # Borde derecho
            top=Side(border_style='thin', color='000000'),    # Borde superior
            bottom=Side(border_style='thin', color='000000'))  # Borde inferior

    # Aplicar formato negrita a los encabezados
    for celda in hoja_trabajo[5]:
        # Configurar el relleno (color de fondo)
        celda.fill = PatternFill(start_color="001F3F", end_color="001F3F", fill_type="solid")  # Código de color para azul marino
        # Configurar el texto (fuente)
        celda.font = Font(color="FFFFFF", bold=True)  # Letras blancas y negritas
        # Configurar la alineación
        celda.alignment = Alignment(horizontal='center')
        # Configurar los bordes
        celda.border = borde

    # Aplicar el borde a las celdas en el rango A-G para las filas específicas
    for i in range(5, len(datos_hr_extra_total) + 5):
        for columna in 'ABCDEF':
            celda = hoja_trabajo[columna + str(i + 1)]
            celda.border = borde
            #hoja_trabajo['B' + str(i+1)] = ''

    for celda in hoja_trabajo[len(datos_hr_extra_total)+8]:
        # Configurar el relleno (color de fondo)
        celda.fill = PatternFill(start_color="001F3F", end_color="001F3F", fill_type="solid")  # Código de color para azul marino
        # Configurar el texto (fuente)
        celda.font = Font(color="FFFFFF", bold=True)  # Letras blancas y negritas
        # Configurar la alineación
        celda.alignment = Alignment(horizontal='center')
        # Configurar los bordes
        celda.border = borde

    # Aplicar el borde a las celdas en el rango A-G para las filas específicas
    for i in range(len(datos_hr_extra_total)+8, len(datos_hr_extra_total) + 8 + len(datos_hr_extra)):
        for columna in 'ABCDEFG':
            celda = hoja_trabajo[columna + str(i + 1)]
            celda.border = borde

    # Configurar el formato para F3 y F2 (fondo azul marino y letras blancas en negrita)
    celda_f2 = hoja_trabajo['F2']
    celda_f2.fill = PatternFill(start_color="001F3F", end_color="001F3F", fill_type="solid")  # Fondo azul marino
    celda_f2.font = Font(color="FFFFFF", bold=True)  # Letras blancas y negritas
    celda_f2.alignment = Alignment(horizontal='center', vertical='center')  # Centrar contenido
    celda_f3 = hoja_trabajo['F3']
    celda_f3.fill = PatternFill(start_color="001F3F", end_color="001F3F", fill_type="solid")  # Fondo azul marino
    celda_f3.font = Font(color="FFFFFF", bold=True)  # Letras blancas y negritas
    celda_f3.alignment = Alignment(horizontal='center', vertical='center')  # Centrar contenido
    celda_g2 = hoja_trabajo['G2']
    celda_g2.border = borde
    celda_g3 = hoja_trabajo['G3']
    celda_g3.border = borde

    celda_titulo = hoja_trabajo['D1']
    celda_titulo.font = Font(name='Arial', size=16, color="000000", bold=True)  # Fuente Arial 16 en negrita y color negro
    celda_titulo.alignment = Alignment(horizontal='center', vertical='center')  # Centrar contenido

    celda_AOp = hoja_trabajo['A' + str(len(datos_hr_extra_total) + 7)]
    celda_AOp.font = Font(name='Arial', size=12, color="000000", bold=True)  # Fuente Arial 16 en negrita y color negro

    celda_firma = hoja_trabajo['C' + str(len(datos_hr_extra_total) + 7 + len(datos_hr_extra) + 5)]
    celda_firma.font = Font(name='Arial', size=12, color="000000", bold=True)  # Fuente Arial 16 en negrita y color negro

    for i in range(5, len(datos_hr_extra_total) + 5):
        if hoja_trabajo['B' + str(i + 1)].value == "06:00am a 06:00 am":
            hoja_trabajo['B' + str(i + 1)]= "N/A"
            hoja_trabajo['C' + str(i + 1)]= "-----"

    for i in range(5, len(datos_hr_extra_total) + 5):
        if hoja_trabajo['D' + str(i + 1)].value == "05:00pm a 05:00 pm" or hoja_trabajo['D' + str(i + 1)].value == "03:00pm a 03:00 pm" :
            hoja_trabajo['D' + str(i + 1)]= "N/A"
            hoja_trabajo['E' + str(i + 1)]= "-----"

    # Configurar el borde a la izquierda en G5
    hoja_trabajo['G5'].fill = PatternFill(start_color=None, end_color=None, fill_type=None)  # Limpiar el formato
    hoja_trabajo['G5'].border = Border(left=Side(border_style="thin", color="000000"))

    # Guardar el archivo y cerrar el libro de trabajo
    workbook.save(nombre_archivo)
    workbook.close()

#Interfaz***************************************************************************************************************************************************************************************
ventana = tk.Tk()
ventana.title("Breton Industrial")
ventana.geometry("1310x815")
#ventana.geometry("1100x475")
ventana.configure(bg="#000040")
#ventana.resizable(False, False)

Logo = PhotoImage(file="Breton_industrial.png")
Logo = Logo.subsample(4, 4)
#Logo = Logo.zoom(4, 4)
etiqueta = tk.Label(ventana, image=Logo)
etiqueta.place(x=15, y=15)

etiqueta_hrstotales = tk.Label(ventana, text="Horas totales", fg="white", font=("Tahoma", 14, "bold"), bg="#000040", width=12, height=1)
etiqueta_hrstotales.place(x=35, y=145)
etiqueta_actividades = tk.Label(ventana, text="Actividades", fg="white", font=("Tahoma", 14, "bold"), bg="#000040", width=12, height=1)
etiqueta_actividades.place(x=35, y=440)
etiqueta_fecha = tk.Label(ventana, text="Fecha", fg="white", font=("Tahoma", 15, "bold"), bg="#000040", width=5, height=1)
etiqueta_fecha.place(x=35, y=95)
etiqueta_imprimir = tk.Label(ventana, text="Imprimir", fg="white", font=("Tahoma", 12, "normal"), bg="#000040", width=7, height=1)
etiqueta_imprimir.place(x=1195, y=775)

date_fecha = DateEntry(ventana, width=10, state="readonly", background='darkblue', foreground='white', borderwidth=3, font=("Tahoma", 12))
date_fecha.place(x=110, y=100)
date_fecha.configure(date_pattern='dd/mm/yyyy')

#text_fecha = tk.Entry(ventana, state="readonly", width=10, font=("Tahoma", 12))
#text_fecha.place(x=155, y=95)
#text_empleado = tk.Entry(ventana, state="readonly", width=25, font=("Tahoma", 12))
#text_empleado.place(x=155, y=130) 

Imagen_imprimir = tk.PhotoImage(file="imprimir.png")
Imprimir = tk.Button(ventana, image=Imagen_imprimir, command=exportar_excel, text="Borrar", width=50, height=50) 
Imprimir.place(x=1200, y=720)

estilo_HrExtraTotal = ttk.Style()
# Asegúrate de que estás utilizando un tema que permita personalizaciones
estilo_HrExtraTotal.theme_use('clam')  # Puedes probar con otros temas si es necesario
# Personalizar el estilo de los encabezados
estilo_HrExtraTotal.configure("Treeview.Heading", background="#808080", foreground="white", font=('Helvetica', 12, 'bold'))
HrExtraTotal = ttk.Treeview(ventana, columns=(0, 1))
HrExtraTotal.place(x=30, y=175)
# 5. Agregar encabezados
HrExtraTotal.heading('#0', text='Nombre')
HrExtraTotal.column('#0', width=250, anchor="w")  # Ancho de la Columna 1
HrExtraTotal.heading('#1', text='Extra Matutino')
HrExtraTotal.column('#1', width=200, anchor="center")  # Ancho de la Columna 2
HrExtraTotal.heading('#2', text='Extra Vespertino')
HrExtraTotal.column('#2', width=175, anchor="center")  # Ancho de la Columna 3

HrExtraTotal.tag_configure("grande", font=('Arial', 11))
scroll_y = ttk.Scrollbar(ventana, orient="vertical", command=HrExtraTotal.yview)
scroll_y.place(x=657, y=205, height=203)
HrExtraTotal.configure(yscrollcommand=scroll_y.set)

estilo_HrExtra = ttk.Style()
# Asegúrate de que estás utilizando un tema que permita personalizaciones
estilo_HrExtra.theme_use('clam')  # Puedes probar con otros temas si es necesario
# Personalizar el estilo de los encabezados
estilo_HrExtra.configure("Treeview.Heading", background="#808080", foreground="white", font=('Helvetica', 12, 'bold'))
HrExtra = ttk.Treeview(ventana, columns=(0, 1, 2, 3, 4, 5))
HrExtra.place(x=30, y=470)
# 5. Agregar encabezados
HrExtra.heading('#0', text='Nombre')
HrExtra.column('#0', width=250, anchor="w")  # Ancho de la Columna 1
HrExtra.heading('#1', text='Horario')
HrExtra.column('#1', width=150, anchor="center")  # Ancho de la Columna 2
HrExtra.heading('#2', text='OT')
HrExtra.column('#2', width=100, anchor="center")  # Ancho de la Columna 3
HrExtra.heading('#3', text='Proyecto')
HrExtra.column('#3', width=300, anchor="center")  # Ancho de la Columna 4
HrExtra.heading('#4', text='Ensamble')
HrExtra.column('#4', width=200, anchor="center")  # Ancho de la Columna 4
HrExtra.heading('#5', text='Operacion')
HrExtra.column('#5', width=150, anchor="center")
HrExtra.heading('#6', text='HorasE')
HrExtra.column('#6', width=80, anchor="center")
HrExtra.tag_configure("grande", font=('Arial', 11))
scroll_y = ttk.Scrollbar(ventana, orient="vertical", command=HrExtra.yview)
scroll_y.place(x=1262, y=500, height=203)
HrExtra.configure(yscrollcommand=scroll_y.set)

#Inicio*****************************************************************************************************************************************************************************************
date_fecha.bind("<<DateEntrySelected>>", ver_HrsExtra)
ver_HrsExtra()

ventana.mainloop()